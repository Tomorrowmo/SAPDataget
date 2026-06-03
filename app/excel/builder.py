"""Excel 生成器 (§10)。

支持的能力 (v0.3):
  * 默认模板:浅色简洁,粗体表头,数字千分位,冻结首行
  * 模板填充 (path A §10.2):Skill 提供 template.xlsx,后端打开模板写入数据区
  * 多 sheet (path C §10.2):一次生成多张 sheet (对比类查询)
  * 敏感字段脱敏 (§14.2):redact / partial / hash 三种 mask 策略
  * 图表嵌入 (F13):从 chart.json 读 chart 配置,用 openpyxl chart API 注入
  * 数据合理性自检 (F11):空值占比 / 异常负数等,写入「查询信息」sheet
"""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import logging
import shutil
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


@dataclass
class ExcelResult:
    path: Path
    rows: int
    columns: list[str]
    size_bytes: int
    sheets: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class SheetSpec:
    """多 sheet 输出时,每个 sheet 的描述。"""
    title: str
    columns: list[str]
    rows: list[dict[str, Any]]
    labels: dict[str, str] | None = None
    chart: dict[str, Any] | None = None


# ============================== 样式常量 ==============================

_HEADER_FILL = PatternFill("solid", fgColor="EEEEEE")
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="333333")
_BODY_FONT = Font(name="微软雅黑", size=10, color="333333")
_THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


# ============================== 脱敏 ==============================


def _mask_value(value: Any, mode: str) -> Any:
    """按 mask_mode 处理单个字段值。"""
    if value is None:
        return value
    s = str(value)
    if not s:
        return s
    if mode == "redact":
        return "***"
    if mode == "partial":
        if len(s) <= 2:
            return "*" * len(s)
        if len(s) <= 4:
            return s[0] + "*" * (len(s) - 1)
        return f"{s[0]}{'*' * (len(s) - 2)}{s[-1]}"
    if mode == "hash":
        h = hashlib.sha256(s.encode("utf-8")).hexdigest()[:8]
        return f"#{h}"
    return value


def apply_sensitive_mask(
    rows: list[dict[str, Any]],
    columns: Iterable[str],
    sensitive_fields: dict[str, str],
) -> tuple[list[dict[str, Any]], list[str]]:
    """对 rows 中匹配 sensitive_fields 的字段批量脱敏,返回新 rows + 实际被 mask 的字段列表。

    sensitive_fields: {field_name: mask_mode}
    """
    if not sensitive_fields:
        return rows, []
    cols = set(columns)
    masked_cols = sorted(c for c in cols if c in sensitive_fields)
    if not masked_cols:
        return rows, []
    out: list[dict[str, Any]] = []
    for r in rows:
        rr = dict(r)
        for c in masked_cols:
            mode = sensitive_fields[c]
            rr[c] = _mask_value(rr.get(c), mode)
        out.append(rr)
    return out, masked_cols


# ============================== Builder ==============================


class TemplateScanError(ValueError):
    """模板上传时检测到不安全内容(VBA 宏等)。"""


def scan_template_safety(path: Path) -> None:
    """扫描 .xlsx/.xlsm,如含 VBA 宏直接拒绝 (§14)。"""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    if p.suffix.lower() == ".xlsm":
        raise TemplateScanError(".xlsm 含宏,拒绝上传 —— 请另存为 .xlsx")
    try:
        with zipfile.ZipFile(p) as zf:
            names = [n.lower() for n in zf.namelist()]
            for n in names:
                if "vbaproject.bin" in n:
                    raise TemplateScanError("检测到 vbaProject.bin (VBA 宏),拒绝")
                if "macrosheets/" in n:
                    raise TemplateScanError("检测到 macrosheets,拒绝")
    except zipfile.BadZipFile as e:
        raise TemplateScanError(f"不是合法 xlsx 文件: {e}") from e


class ExcelBuilder:
    """openpyxl 简易封装。无外部状态,可重复使用。"""

    def __init__(self, output_dir: Path) -> None:
        self.output_dir = Path(output_dir).resolve()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def build(
        self,
        *,
        filename: str,
        columns: list[str],
        rows: list[dict[str, Any]],
        labels: dict[str, str] | None = None,
        info: dict[str, Any] | None = None,
        sheet_name: str = "数据",
        sensitive_fields: dict[str, str] | None = None,
        template_path: Path | None = None,
        chart: dict[str, Any] | None = None,
    ) -> ExcelResult:
        """单 sheet 生成。

        Args:
            sensitive_fields: {field_name: mask_mode} —— 命中则写入前先 mask。
            template_path: 若提供且 .xlsx 存在,以模板为基底打开;主 sheet 名取
                          模板的第一张 sheet 或 ``sheet_name`` 匹配的那张。
            chart: {kind: bar|line|pie, x: 列名, y: [列名...], title: ...} —— 在
                   主 sheet 上嵌入图表。
        """
        return self.build_multi(
            filename=filename,
            sheets=[SheetSpec(
                title=sheet_name,
                columns=columns,
                rows=rows,
                labels=labels,
                chart=chart,
            )],
            info=info,
            sensitive_fields=sensitive_fields,
            template_path=template_path,
        )

    def build_multi(
        self,
        *,
        filename: str,
        sheets: list[SheetSpec],
        info: dict[str, Any] | None = None,
        sensitive_fields: dict[str, str] | None = None,
        template_path: Path | None = None,
    ) -> ExcelResult:
        """多 sheet 生成 (F12)。第一张 sheet 视为主数据。"""
        info = dict(info or {})
        sensitive_fields = sensitive_fields or {}

        all_warnings: list[str] = []
        all_masked: set[str] = set()
        processed: list[SheetSpec] = []
        for sp in sheets:
            masked_rows, masked_cols = apply_sensitive_mask(sp.rows, sp.columns, sensitive_fields)
            if masked_cols:
                all_masked.update(masked_cols)
            warns = sanity_check(masked_rows, sp.columns)
            all_warnings.extend(f"[{sp.title}] {w}" for w in warns)
            processed.append(SheetSpec(
                title=sp.title,
                columns=sp.columns,
                rows=masked_rows,
                labels=sp.labels,
                chart=sp.chart,
            ))

        # 模板或全新
        if template_path and Path(template_path).exists():
            wb = load_workbook(template_path)
            # 清空模板里 (除"查询信息"外) 的现有数据,保留样式靠模板提供
            for ws in list(wb.worksheets):
                if ws.title == "查询信息":
                    wb.remove(ws)
            used_template = True
        else:
            wb = Workbook()
            # 移除默认 Sheet1
            for ws in list(wb.worksheets):
                wb.remove(ws)
            used_template = False

        # 写每张数据 sheet
        for idx, sp in enumerate(processed):
            target_ws = None
            if used_template and idx == 0 and wb.worksheets:
                # 把模板的第一张当主数据 sheet,改名为 sp.title
                target_ws = wb.worksheets[0]
                target_ws.title = sp.title
                _clear_sheet_data(target_ws)
            if target_ws is None:
                target_ws = wb.create_sheet(sp.title)
            _write_data_sheet(target_ws, sp.columns, sp.rows, sp.labels or {})
            if sp.chart:
                try:
                    _embed_chart(target_ws, sp.columns, sp.rows, sp.chart)
                except Exception as e:                                # noqa: BLE001
                    log.warning("图表嵌入失败 %s: %s", sp.title, e)
                    all_warnings.append(f"[{sp.title}] 图表失败: {e}")

        # 查询信息 sheet
        if all_masked:
            info["masked_fields"] = sorted(all_masked)
        if all_warnings:
            info["sanity_warnings"] = all_warnings
        meta_ws = wb.create_sheet("查询信息")
        _write_info_sheet(meta_ws, info)

        # 保存
        out_path = self.output_dir / filename
        wb.save(out_path)
        size = out_path.stat().st_size

        primary = processed[0] if processed else None
        return ExcelResult(
            path=out_path,
            rows=len(primary.rows) if primary else 0,
            columns=list(primary.columns) if primary else [],
            size_bytes=size,
            sheets=[w.title for w in wb.worksheets],
            warnings=all_warnings,
        )


# ============================== sheet writers ==============================


def _clear_sheet_data(ws: Any) -> None:
    """清空 worksheet 所有单元格但保留样式 (用于模板填充)。"""
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
        return
    ws.delete_rows(1, ws.max_row)


def _write_data_sheet(
    ws: Any,
    columns: list[str],
    rows: list[dict[str, Any]],
    labels: dict[str, str],
) -> None:
    if not columns:
        ws.cell(row=1, column=1, value="(无数据)")
        return

    # 表头
    for j, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=j, value=labels.get(col, col))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    # 数据
    for i, r in enumerate(rows, start=2):
        for j, col in enumerate(columns, start=1):
            val = r.get(col)
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = "#,##0.##" if isinstance(val, float) else "#,##0"
                cell.alignment = _RIGHT
            else:
                cell.alignment = _LEFT

    # 列宽自适应
    for j, col in enumerate(columns, start=1):
        max_len = len(str(labels.get(col, col)))
        for r in rows[:200]:
            v = r.get(col)
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(40, max(8, max_len + 2))

    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 100


def _write_info_sheet(ws: Any, info: dict[str, Any]) -> None:
    rows: list[tuple[str, Any]] = [
        ("生成时间", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    known_order = [
        "username", "question", "skill_id", "skill_version",
        "service", "entity_set", "odata_url",
        "row_count", "latency_ms",
        "llm_model", "llm_input_tokens", "llm_output_tokens",
        "bw_mode", "masked_fields", "sanity_warnings",
    ]
    seen: set[str] = set()
    for k in known_order:
        if k in info:
            rows.append((k, info[k]))
            seen.add(k)
    for k, v in info.items():
        if k not in seen:
            rows.append((k, v))

    ws.cell(row=1, column=1, value="字段").font = _HEADER_FONT
    ws.cell(row=1, column=2, value="值").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=2).fill = _HEADER_FILL

    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=str(k)).font = _BODY_FONT
        cell = ws.cell(row=i, column=2, value=_stringify(v))
        cell.font = _BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80


def _stringify(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list, tuple)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


# ============================== Sanity check (F11) ==============================


def sanity_check(rows: list[dict[str, Any]], columns: list[str]) -> list[str]:
    """对数据做基础合理性检查,返回告警列表 (可能为空)。"""
    warns: list[str] = []
    if not rows:
        return warns
    n = len(rows)
    for c in columns:
        nulls = sum(1 for r in rows if r.get(c) in (None, "", "null"))
        if nulls / n > 0.5:
            warns.append(f"列 {c} 空值占比 {nulls / n:.0%},数据可能不完整")
        # 数值列异常负数检查 (仅对名字看起来是金额/数量的列)
        if any(tok in c.upper() for tok in ("AMOUNT", "NETWR", "QTY", "STOCK", "PROFIT")):
            negs = sum(1 for r in rows if isinstance(r.get(c), (int, float))
                       and not isinstance(r.get(c), bool) and r.get(c) < 0)
            if negs > 0 and negs / n > 0.3:
                warns.append(f"列 {c} 负数占比 {negs / n:.0%},请确认")
    return warns


# ============================== 图表嵌入 (F13) ==============================


def _embed_chart(
    ws: Any,
    columns: list[str],
    rows: list[dict[str, Any]],
    chart_cfg: dict[str, Any],
) -> None:
    """根据 chart_cfg 在 ws 上嵌入图表。

    chart_cfg 形如:
        {"kind": "bar"|"line"|"pie", "x": "OfficeCode",
         "y": ["NETWR_F", "GROSS_PROFIT"], "title": "..."}
    """
    if not rows or not columns:
        return
    kind = (chart_cfg.get("kind") or "bar").lower()
    x_col = chart_cfg.get("x")
    y_cols_raw = chart_cfg.get("y") or []
    if isinstance(y_cols_raw, str):
        y_cols = [y_cols_raw]
    else:
        y_cols = list(y_cols_raw)
    if not x_col or not y_cols:
        return
    if x_col not in columns:
        return
    y_cols = [c for c in y_cols if c in columns]
    if not y_cols:
        return

    x_col_idx = columns.index(x_col) + 1
    n_rows = len(rows)
    data_start_row = 2
    data_end_row = data_start_row + n_rows - 1

    if kind == "line":
        chart = LineChart()
    elif kind == "pie":
        chart = PieChart()
    else:
        chart = BarChart()
    chart.title = chart_cfg.get("title") or ""
    chart.x_axis.title = x_col
    chart.y_axis.title = ", ".join(y_cols)
    chart.style = 11
    chart.height = 10
    chart.width = 20

    # X 轴 categories
    cats = Reference(ws, min_col=x_col_idx, min_row=data_start_row,
                     max_row=data_end_row)
    for y in y_cols:
        y_idx = columns.index(y) + 1
        data_ref = Reference(ws, min_col=y_idx, min_row=1,
                             max_row=data_end_row, max_col=y_idx)
        chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    # 放在数据右侧
    anchor_col_letter = get_column_letter(len(columns) + 2)
    ws.add_chart(chart, f"{anchor_col_letter}2")


# ============================== utility: load chart cfg ==============================


def load_chart_config(skill_folder: Path | None) -> dict[str, Any] | None:
    """从 skill 文件夹读 chart.json (若有)。"""
    if not skill_folder:
        return None
    p = Path(skill_folder) / "chart.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:                                            # noqa: BLE001
        log.warning("chart.json 解析失败 %s: %s", p, e)
        return None
