"""启动复检脚本 —— 把后端关键端点跑一遍,给出状态报告。

用法:
    python -m cli.health
"""
from __future__ import annotations

import io
import json
import sys
from typing import Any

import httpx
from openpyxl import load_workbook

if sys.platform == "win32":
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")              # type: ignore[attr-defined]
        except (AttributeError, ValueError):
            pass

BASE = "http://127.0.0.1:8000"


def main() -> int:
    try:
        # trust_env=False —— 关掉系统 HTTP_PROXY,避免代理把 127.0.0.1 也吃了
        c = httpx.Client(base_url=BASE, timeout=30.0, trust_env=False)
    except Exception as e:
        print(f"❌ 无法连接: {e}")
        return 1

    print("=" * 64)
    print(" SAP BW 智能取数平台 — 启动复检")
    print("=" * 64)

    # 1. 状态
    s = c.get("/api/status").json()
    print(f"\n[1/5] /api/status")
    print(f"  版本:     {s['version']}")
    print(f"  BW 模式:  {s['bw_mode']}  ({s['bw']})")
    print(f"  当前 LLM: {s['llm']['current_display']}")
    print(f"           {s['llm']['current']}  ready={s['llm']['current_ready']}")
    print(f"  可选模型: {len(s['llm']['models'])} 个")
    print(f"  Skills:   {s['skills_count']} 个")

    # 2. 登录(admin)
    print(f"\n[2/5] POST /api/auth/login (admin)")
    r = c.post("/api/auth/login", json={"username": "admin", "password": ""})
    if r.status_code != 200:
        print(f"  ❌ 登录失败: {r.status_code} {r.text}")
        return 1
    me = r.json()
    print(f"  ✅ {me['username']} / {me['role']}")

    # 3. Skills
    print(f"\n[3/5] GET /api/skills")
    skills = c.get("/api/skills").json()["skills"]
    for sk in skills:
        print(f"  📊 {sk['id']:<24} {sk['title']}")

    # 4. 跑一个 Skill 验证 Excel 落盘
    print(f"\n[4/5] POST /api/skills/monthly_sales_region/run")
    r = c.post(
        "/api/skills/monthly_sales_region/run",
        json={"params": {"month": "202605", "region": "HD", "top_n": 5}},
    )
    if r.status_code != 200:
        print(f"  ❌ {r.status_code} {r.text}")
        return 1
    data = r.json()
    print(f"  状态:     {data['status']}")
    print(f"  行数:     {data['row_count']}")
    if data["excel"]:
        print(f"  Excel:    {data['excel']['filename']}  ({data['excel']['size_bytes']:,} bytes)")
        # 下载验证
        dr = c.get(data["excel"]["download_url"])
        if dr.status_code == 200:
            wb = load_workbook(io.BytesIO(dr.content))
            print(f"  下载校验: sheets={wb.sheetnames}")

    # 5. 任务历史
    print(f"\n[5/5] GET /api/tasks")
    tasks = c.get("/api/tasks").json()
    print(f"  历史任务总数: {tasks['total']}")
    for t in tasks["tasks"][:5]:
        marker = "✅" if t["status"] == "done" else "❌"
        print(f"  {marker} {t['created_at']}  {t['source']:<6} {(t['skill_id'] or '-'):<24} {t['row_count'] or 0} 行")

    print("\n" + "=" * 64)
    print(" ✅ 启动复检通过 — 系统就绪")
    print("=" * 64)
    print(f"\n 🌐 在浏览器打开: {BASE}")
    print("    - 用 admin 登录获得管理员权限,可访问 Skill 管理 / 审计日志 / 敏感字段")
    print("    - 用任意其他用户名登录获得业务用户角色")
    print(f"\n 📖 API 文档:    {BASE}/docs (Swagger UI)\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
