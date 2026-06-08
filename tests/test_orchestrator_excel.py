"""Orchestrator + ExcelBuilder 端到端测试 (不接 LLM)。"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.bw.mock import MockBWClient
from app.config import BWSettings, LLMSettings, Settings
from app.orchestrator import TaskOrchestrator
from app.skills.registry import SkillRegistry


@pytest.fixture
def settings(tmp_path: Path, mock_data_dir: Path) -> Settings:
    return Settings(
        llm=LLMSettings(model="dummy/none", api_base=None, api_key=None, timeout=10, max_iters=3),
        bw=BWSettings(
            mode="mock",
            mock_data_dir=mock_data_dir,
            mock_latency_ms=0,
            base_url="", username="", password="", client="",
            language="EN", verify_ssl=True, timeout=60,
            client_fallback=True, max_export_rows=50000,
        ),
        output_dir=tmp_path / "outputs",
        skills_dir=Path(__file__).resolve().parent.parent / "data" / "skills",
        owner_field="UName",
    )


@pytest.fixture
def orch(settings: Settings, mock_bw: MockBWClient) -> TaskOrchestrator:
    skills = SkillRegistry(settings.skills_dir)
    skills.reload()
    return TaskOrchestrator(settings, mock_bw, skills)


def test_run_skill_produces_excel(orch: TaskOrchestrator, tmp_path: Path):
    result = orch.run_skill(
        "monthly_sales_region",
        {"month": "202605", "region": "HD", "top_n": 5},
        username="alice",
    )
    assert result.status == "done", result.error
    assert result.excel is not None
    assert result.excel.path.exists()
    assert result.excel.size_bytes > 1000          # 真的写了内容
    assert result.row_count >= 1

    # 验证 xlsx 内容
    wb = load_workbook(result.excel.path)
    assert "查询信息" in wb.sheetnames
    data_ws = wb[wb.sheetnames[0]]
    # 表头
    headers = [c.value for c in data_ws[1]]
    assert any(h and "销售" in str(h) for h in headers), f"应有中文 label: {headers}"
    # 行数
    data_rows = list(data_ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == result.row_count

    # 查询信息 sheet
    info_ws = wb["查询信息"]
    info_dict = {}
    for r in info_ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            info_dict[r[0]] = r[1]
    assert info_dict.get("username") == "alice"
    assert info_dict.get("skill_id") == "monthly_sales_region"
    assert info_dict.get("bw_mode") == "mock"


def test_run_free_query_scopes_rows_by_owner(orch: TaskOrchestrator):
    """run_free_query:结果含 UName 时只保留登录用户名下的行(Excel + 预览 + 计数)。"""
    rows = [
        {"UName": "ADMIN", "Val": 1},
        {"UName": "BOB", "Val": 2},
        {"UName": "admin", "Val": 3},   # 大小写归一化后仍属 admin
    ]
    res = orch.run_free_query(
        service="ZX_SRV", entity_set="Items", columns=["UName", "Val"],
        rows=rows, info={"row_count": 3}, username="admin",
    )
    assert res.status == "done"
    assert res.row_count == 2, "应只含 admin 的 2 行"
    assert {r["Val"] for r in res.rows_preview} == {1, 3}
    assert res.meta.get("owner_scoped") is True

    # 不含 UName 的结果不受影响(no-op)
    res2 = orch.run_free_query(
        service="ZX_SRV", entity_set="Items", columns=["Region", "Val"],
        rows=[{"Region": "HD", "Val": 9}], info={"row_count": 1}, username="admin",
    )
    assert res2.row_count == 1
    assert res2.meta.get("owner_scoped") is not True


def test_run_skill_unknown_id(orch: TaskOrchestrator):
    result = orch.run_skill("nosuch", {}, username="bob")
    assert result.status == "failed"
    assert "Skill" in (result.error or "")


def test_run_skill_empty_result(orch: TaskOrchestrator):
    """查询条件不匹配任何数据 → 友好失败。"""
    result = orch.run_skill(
        "monthly_sales_region",
        {"month": "190001", "region": "HD"},     # 1900-01 mock 里没有
        username="bob",
    )
    assert result.status == "failed"
    assert result.error is not None
