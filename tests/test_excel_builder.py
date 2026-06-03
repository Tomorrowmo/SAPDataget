"""ExcelBuilder 直接测试 —— 脱敏、模板填充、图表、多 sheet、安全扫描。"""
from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.excel.builder import (
    ExcelBuilder,
    SheetSpec,
    TemplateScanError,
    _mask_value,
    apply_sensitive_mask,
    load_chart_config,
    sanity_check,
    scan_template_safety,
)


# ============================== mask 单元 ==============================


def test_mask_redact_replaces_with_asterisks():
    assert _mask_value("hello", "redact") == "***"


def test_mask_partial_short_string():
    assert _mask_value("ab", "partial") == "**"
    assert _mask_value("abc", "partial") == "a**"


def test_mask_partial_long_string():
    assert _mask_value("zhang3@example.com", "partial") == "z****************m"


def test_mask_hash_deterministic_and_prefixed():
    a = _mask_value("zhang3", "hash")
    b = _mask_value("zhang3", "hash")
    assert a == b and a.startswith("#") and len(a) == 9    # # + 8 hex


def test_mask_none_value_pass_through():
    assert _mask_value(None, "redact") is None


def test_mask_unknown_mode_pass_through():
    assert _mask_value("x", "bogus") == "x"


def test_apply_sensitive_mask_filters_to_present_columns():
    rows = [{"a": "secret", "b": 100}, {"a": "another", "b": 200}]
    masked, cols = apply_sensitive_mask(rows, ["a", "b"], {"a": "redact", "c": "redact"})
    assert cols == ["a"]
    assert masked[0]["a"] == "***" and masked[0]["b"] == 100


def test_apply_sensitive_mask_empty_when_no_match():
    rows = [{"x": 1}]
    masked, cols = apply_sensitive_mask(rows, ["x"], {"y": "redact"})
    assert masked == rows
    assert cols == []


# ============================== sanity check ==============================


def test_sanity_check_flags_high_null_ratio():
    rows = [{"a": None}, {"a": None}, {"a": None}, {"a": 1}]
    warns = sanity_check(rows, ["a"])
    assert any("空值" in w for w in warns)


def test_sanity_check_flags_money_negatives():
    rows = [{"NETWR_F": -10}] * 7 + [{"NETWR_F": 1}] * 3
    warns = sanity_check(rows, ["NETWR_F"])
    assert any("负数" in w for w in warns)


def test_sanity_check_empty_rows_no_warnings():
    assert sanity_check([], ["a"]) == []


# ============================== Build basic / masking / multi / chart ==============================


def test_build_basic_writes_xlsx_with_two_sheets(tmp_path: Path):
    b = ExcelBuilder(tmp_path)
    res = b.build(
        filename="t1.xlsx",
        columns=["x", "y"],
        rows=[{"x": 1, "y": 2.5}, {"x": 3, "y": 4.5}],
        info={"username": "alice", "service": "S"},
    )
    assert res.path.exists() and res.rows == 2
    wb = load_workbook(res.path)
    assert "数据" in wb.sheetnames and "查询信息" in wb.sheetnames


def test_build_with_sensitive_mask_writes_masked_value_and_records_in_info(tmp_path: Path):
    b = ExcelBuilder(tmp_path)
    res = b.build(
        filename="masked.xlsx",
        columns=["NAME", "SALARY"],
        rows=[{"NAME": "张三", "SALARY": 12345}, {"NAME": "李四", "SALARY": 23000}],
        info={"service": "HR"},
        sensitive_fields={"SALARY": "redact"},
    )
    wb = load_workbook(res.path)
    ws = wb["数据"]
    assert ws.cell(2, 2).value == "***"
    assert ws.cell(3, 2).value == "***"
    info_ws = wb["查询信息"]
    info = {r[0]: r[1] for r in info_ws.iter_rows(min_row=2, values_only=True) if r[0]}
    assert "SALARY" in str(info.get("masked_fields", ""))


def test_build_multi_creates_one_sheet_per_spec(tmp_path: Path):
    b = ExcelBuilder(tmp_path)
    sheets = [
        SheetSpec(title="Q1", columns=["a"], rows=[{"a": 1}]),
        SheetSpec(title="Q2", columns=["a"], rows=[{"a": 2}]),
        SheetSpec(title="Q3", columns=["a"], rows=[{"a": 3}]),
    ]
    res = b.build_multi(filename="multi.xlsx", sheets=sheets, info={})
    wb = load_workbook(res.path)
    assert {"Q1", "Q2", "Q3", "查询信息"}.issubset(set(wb.sheetnames))


def test_build_with_chart_embeds_chart(tmp_path: Path):
    b = ExcelBuilder(tmp_path)
    res = b.build(
        filename="chart.xlsx",
        columns=["Office", "Sales"],
        rows=[{"Office": "A", "Sales": 100}, {"Office": "B", "Sales": 200}],
        info={},
        chart={"kind": "bar", "x": "Office", "y": ["Sales"], "title": "Test"},
    )
    wb = load_workbook(res.path)
    ws = wb["数据"]
    assert len(ws._charts) == 1


def test_build_with_template_uses_template_as_base(tmp_path: Path):
    # 构造一个 template.xlsx,标个标识单元格
    tpl = tmp_path / "tpl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.cell(10, 10).value = "TEMPLATE_MARKER"   # 占位
    wb.save(tpl)

    b = ExcelBuilder(tmp_path)
    res = b.build(
        filename="from_tpl.xlsx",
        columns=["x"],
        rows=[{"x": 1}],
        info={},
        sheet_name="数据",
        template_path=tpl,
    )
    wb2 = load_workbook(res.path)
    ws2 = wb2["数据"]
    # 模板写入会清空数据 sheet (保留样式),所以 marker 不应保留
    # 主要验证从模板打开能正常保存
    assert ws2.cell(1, 1).value == "x"   # 表头写上了
    assert ws2.cell(2, 1).value == 1


# ============================== Template safety scan ==============================


def test_scan_template_rejects_xlsm(tmp_path: Path):
    p = tmp_path / "x.xlsm"
    Workbook().save(p)
    with pytest.raises(TemplateScanError):
        scan_template_safety(p)


def test_scan_template_rejects_xlsx_with_vbaProject(tmp_path: Path):
    # 伪造一个含 vbaProject.bin 的 zip
    p = tmp_path / "evil.xlsx"
    with zipfile.ZipFile(p, "w") as zf:
        zf.writestr("xl/vbaProject.bin", b"BOGUS")
        zf.writestr("[Content_Types].xml", "<x/>")
    with pytest.raises(TemplateScanError, match="VBA"):
        scan_template_safety(p)


def test_scan_template_accepts_plain_xlsx(tmp_path: Path):
    p = tmp_path / "ok.xlsx"
    Workbook().save(p)
    scan_template_safety(p)              # 不抛即通过


def test_scan_template_rejects_non_zip(tmp_path: Path):
    p = tmp_path / "bad.xlsx"
    p.write_bytes(b"not a zip at all")
    with pytest.raises(TemplateScanError):
        scan_template_safety(p)


# ============================== chart.json loader ==============================


def test_load_chart_config_none_when_folder_missing(tmp_path: Path):
    assert load_chart_config(tmp_path / "nonexistent") is None


def test_load_chart_config_returns_dict_when_present(tmp_path: Path):
    cfg = {"kind": "line", "x": "Month", "y": ["NETWR_F"]}
    (tmp_path / "chart.json").write_text(json.dumps(cfg), encoding="utf-8")
    assert load_chart_config(tmp_path) == cfg


def test_load_chart_config_returns_none_on_bad_json(tmp_path: Path):
    (tmp_path / "chart.json").write_text("not json {", encoding="utf-8")
    assert load_chart_config(tmp_path) is None
