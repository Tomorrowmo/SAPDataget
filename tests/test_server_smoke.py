"""端到端 HTTP 烟测 —— 用 httpx 直接打 uvicorn 起的进程。

需要先启动:
    uvicorn app.server:app --port 8000

测试覆盖业务用户主流程:
  login → list-skills → run-skill → download xlsx → list-tasks → switch model
"""
from __future__ import annotations

import io
import os
import time

import httpx
import pytest
from openpyxl import load_workbook

BASE_URL = os.environ.get("BW_TEST_URL", "http://127.0.0.1:8000")


def _delete_all_tasks(client: httpx.Client) -> None:
    tasks = client.get("/api/tasks").json().get("tasks", [])
    for task in tasks:
        client.delete(f"/api/tasks/{task['id']}")


def _server_up() -> bool:
    try:
        r = httpx.get(f"{BASE_URL}/api/status", timeout=2.0)
        return r.status_code == 200
    except Exception:
        return False


pytestmark = pytest.mark.skipif(not _server_up(), reason="uvicorn 未启动")


@pytest.fixture
def client():
    # Windows 上 httpx keep-alive 偶尔被远端强关 → 配置传输层重试
    transport = httpx.HTTPTransport(retries=3)
    with httpx.Client(base_url=BASE_URL, timeout=30.0, transport=transport) as c:
        yield c


@pytest.fixture
def logged_in(client: httpx.Client):
    r = client.post("/api/auth/login", json={"username": "admin", "password": ""})
    assert r.status_code == 200, r.text
    yield client
    _delete_all_tasks(client)


def test_status(client: httpx.Client):
    r = client.get("/api/status")
    assert r.status_code == 200
    data = r.json()
    assert data["bw_mode"] == "mock"
    assert data["skills_count"] >= 3
    # DataAgent 式设置后,llm 块改为 current/current_ready(models 已弃用)
    assert "current" in data["llm"] and "current_ready" in data["llm"]


def test_unauthorized(client: httpx.Client):
    r = client.get("/api/skills")
    assert r.status_code == 401


def test_login_logout(client: httpx.Client):
    r = client.post("/api/auth/login", json={"username": "alice", "password": ""})
    assert r.status_code == 200
    me = client.get("/api/auth/me").json()
    assert me["username"] == "alice"

    r = client.post("/api/auth/logout")
    assert r.status_code == 200
    r = client.get("/api/auth/me")
    assert r.status_code == 401


def test_admin_role(client: httpx.Client):
    """登录 admin 用户应自动获得 admin role (mock 模式约定)"""
    r = client.post("/api/auth/login", json={"username": "admin", "password": ""})
    assert r.status_code == 200
    assert r.json()["role"] == "admin"


def test_skills_list_detail(logged_in: httpx.Client):
    r = logged_in.get("/api/skills")
    assert r.status_code == 200
    data = r.json()
    ids = {s["id"] for s in data["skills"]}
    assert {"monthly_sales_region", "top_customers", "plant_yield"}.issubset(ids)

    r = logged_in.get("/api/skills/monthly_sales_region")
    assert r.status_code == 200
    detail = r.json()
    assert detail["service"] == "ZBW_SALES_SRV"
    pnames = {p["name"] for p in detail["params"]}
    assert {"month", "region"}.issubset(pnames)


def test_run_skill_and_download(logged_in: httpx.Client):
    r = logged_in.post(
        "/api/skills/monthly_sales_region/run",
        json={"params": {"month": "202605", "region": "HD", "top_n": 5}},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["status"] == "done"
    assert data["row_count"] >= 1
    assert data["excel"] is not None
    download_url = data["excel"]["download_url"]
    task_id = data["task_id"]

    # 下载 Excel
    r2 = logged_in.get(download_url)
    assert r2.status_code == 200
    assert r2.headers["content-type"].startswith("application/")
    wb = load_workbook(io.BytesIO(r2.content))
    assert "查询信息" in wb.sheetnames

    # 任务历史里能查到
    tasks = logged_in.get("/api/tasks").json()["tasks"]
    assert any(t["id"] == task_id for t in tasks)


def test_run_skill_validates(logged_in: httpx.Client):
    """缺必填参数应失败 (200 + status=failed,不要 5xx)"""
    r = logged_in.post(
        "/api/skills/monthly_sales_region/run",
        json={"params": {"region": "HD"}},  # 缺 month
    )
    assert r.status_code == 200
    assert r.json()["status"] == "failed"


def test_services_list(logged_in: httpx.Client):
    r = logged_in.get("/api/services")
    assert r.status_code == 200
    names = {s["TechnicalServiceName"] for s in r.json()["services"]}
    assert "ZBW_SALES_SRV" in names


def test_service_metadata(logged_in: httpx.Client):
    r = logged_in.get("/api/services/ZBW_SALES_SRV")
    assert r.status_code == 200
    es = {e["name"] for e in r.json()["entity_sets"]}
    assert "SalesByOfficeView" in es


def test_llm_settings_get(logged_in: httpx.Client):
    """DataAgent 式三元组设置:GET 返回 has_key/model/effective + 模型示例。"""
    r = logged_in.get("/api/llm/settings")
    assert r.status_code == 200, r.text
    body = r.json()
    assert "has_key" in body and "effective_model" in body and "effective_ready" in body
    assert isinstance(body["suggestions"], list) and len(body["suggestions"]) >= 5


def test_admin_endpoints_require_admin(client: httpx.Client):
    # 普通用户登录
    client.post("/api/auth/login", json={"username": "bob", "password": ""})
    assert client.get("/api/audit").status_code == 403
    assert client.get("/api/sensitive-fields").status_code == 403


def test_admin_audit_and_sensitive(logged_in: httpx.Client):
    # admin 已登录
    r = logged_in.get("/api/audit?limit=10")
    assert r.status_code == 200
    assert "audit" in r.json()

    # 敏感字段 CRUD
    r2 = logged_in.post(
        "/api/sensitive-fields",
        json={"service": "ZBW_HR_SRV", "field": "SALARY_BASE", "mask_mode": "redact"},
    )
    assert r2.status_code == 200
    listed = logged_in.get("/api/sensitive-fields").json()["fields"]
    assert any(f["field"] == "SALARY_BASE" for f in listed)
    logged_in.delete("/api/sensitive-fields/ZBW_HR_SRV/SALARY_BASE")
    listed2 = logged_in.get("/api/sensitive-fields").json()["fields"]
    assert not any(f["field"] == "SALARY_BASE" for f in listed2)


def test_llm_settings_per_user_isolation(client: httpx.Client):
    """alice 的设置不应影响 bob —— 每用户独立(DataAgent 式三元组)。"""
    # alice 配三元组
    client.post("/api/auth/login", json={"username": "alice", "password": ""})
    r = client.put("/api/llm/settings", json={
        "api_key": "sk-alice-key-xxxxxxxx",
        "base_url": "https://api.deepseek.com/v1",
        "model": "deepseek-chat",
    })
    assert r.status_code == 200, r.text
    a = r.json()
    assert a["has_key"] is True and a["model"] == "deepseek-chat" and a["effective_ready"] is True

    # bob 登录,不应继承 alice 的设置
    client.post("/api/auth/login", json={"username": "bob", "password": ""})
    b = client.get("/api/llm/settings").json()
    assert b["has_key"] is False, "bob 不应继承 alice 的 key"
    assert b["model"] == "", "bob 的 model 应为空(回退 .env)"

    # 清理
    client.post("/api/auth/login", json={"username": "alice", "password": ""})
    client.put("/api/llm/settings", json={"api_key": "", "base_url": "", "model": ""})


def test_llm_settings_save_keep_and_clear_key(logged_in: httpx.Client):
    """api_key 语义:非空更新 / null 保持 / 空串清空。"""
    # 设 key + model
    r = logged_in.put("/api/llm/settings", json={
        "api_key": "sk-fake-1234567890abcdef", "base_url": "", "model": "deepseek/deepseek-chat",
    })
    assert r.status_code == 200 and r.json()["has_key"] is True

    # 只改 model,api_key=null → key 应保持
    r2 = logged_in.put("/api/llm/settings", json={
        "api_key": None, "base_url": "", "model": "dashscope/qwen-plus",
    })
    body2 = r2.json()
    assert body2["has_key"] is True, "api_key=null 应保持原 key"
    assert body2["model"] == "dashscope/qwen-plus"

    # api_key="" → 清空
    r3 = logged_in.put("/api/llm/settings", json={"api_key": "", "base_url": "", "model": ""})
    assert r3.json()["has_key"] is False, "api_key='' 应清空 key"


def test_report_list_shortcut_works_without_llm(logged_in: httpx.Client):
    r = logged_in.post("/api/chat", json={"message": "报告清单"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["answer"].startswith("已查询到报告清单")
    assert body["llm_model"] == "builtin/report-list"
    assert body["task"]["status"] == "done"
    assert body["task"]["row_count"] >= 1
    assert body["task"]["excel"] is not None
    assert body["task"]["excel"]["download_url"].startswith("/api/tasks/")
    first = body["task"]["rows_preview"][0]
    assert {"ReportID", "ReportDescription"}.issubset(first.keys())


def test_report_list_shortcut_respects_top_n(logged_in: httpx.Client):
    r = logged_in.post("/api/chat", json={"message": "报告清单前1条"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["task"]["status"] == "done"
    assert len(body["task"]["rows_preview"]) == 1


def test_chat_sessions_list_rename_delete(logged_in: httpx.Client):
    """会话历史:用报告清单(无需 LLM)产生 chat 会话,验证 列表/重命名/删除。"""
    r = logged_in.post("/api/chat", json={"message": "报告清单"})
    assert r.status_code == 200, r.text
    tid = r.json()["task_id"]

    sess = logged_in.get("/api/chat/sessions").json()["sessions"]
    assert any(s["id"] == tid for s in sess), "新会话应出现在列表"

    rn = logged_in.patch(f"/api/chat/sessions/{tid}", json={"title": "我的测试会话"})
    assert rn.status_code == 200 and rn.json()["title"] == "我的测试会话"
    sess2 = logged_in.get("/api/chat/sessions").json()["sessions"]
    assert any(s["id"] == tid and s["title"] == "我的测试会话" for s in sess2)

    d = logged_in.delete(f"/api/tasks/{tid}")
    assert d.status_code == 200
    sess3 = logged_in.get("/api/chat/sessions").json()["sessions"]
    assert not any(s["id"] == tid for s in sess3), "删除后不应再在列表"


def test_chat_sessions_per_user_isolation(client: httpx.Client):
    """alice 的会话 bob 看不到、也不能重命名。"""
    client.post("/api/auth/login", json={"username": "alice", "password": ""})
    tid = client.post("/api/chat", json={"message": "报告清单"}).json()["task_id"]

    client.post("/api/auth/login", json={"username": "bob", "password": ""})
    bob_sess = client.get("/api/chat/sessions").json()["sessions"]
    assert not any(s["id"] == tid for s in bob_sess), "bob 不应看到 alice 的会话"
    rn = client.patch(f"/api/chat/sessions/{tid}", json={"title": "hack"})
    assert rn.status_code == 404, "bob 不能重命名 alice 的会话"

    client.post("/api/auth/login", json={"username": "alice", "password": ""})
    client.delete(f"/api/tasks/{tid}")


def test_spa_fallback(client: httpx.Client):
    """前端路由 fallback 应返回 index.html"""
    r = client.get("/some/spa/path")
    assert r.status_code == 200
    assert "<!doctype html>" in r.text.lower() or "<html" in r.text.lower()
